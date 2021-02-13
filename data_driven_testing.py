import openpyxl
from selenium import webdriver

# reading data from excel
file_path = "Barcodes.xlsx"
workbook = openpyxl.load_workbook(file_path)
sheet_names = workbook.sheetnames
sheet = workbook[sheet_names[0]]

rows = sheet.max_row
cols = sheet.max_column

print(rows, cols)

# just use pandas, to read data from excel, to test cases using
# that data, and write back the results
